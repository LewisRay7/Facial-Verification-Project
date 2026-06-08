from __future__ import annotations

import json
import csv
from io import BytesIO, StringIO
from datetime import datetime
from typing import Annotated

from fastapi import APIRouter, Depends, HTTPException, UploadFile, status
from openpyxl import load_workbook
from sqlalchemy.orm import Session
from sqlalchemy import update

from backend.auth.security import require_roles
from backend.database import get_db
from backend.logs.audit import log_event
from backend.security.data_encryption import hash_student_identifier
from backend.models.schemas import (
    EligibleStudentAdd,
    ExamEntryEvaluateIn,
    ExamSessionCreate,
    ExamSessionUpdate,
    InvigilatorAssignmentIn,
)
from backend.models.tables import (
    Device,
    ExamSession,
    ExamSessionInvigilator,
    ExamSessionStudent,
    ExamImportAudit,
    Student,
    User,
    VerificationLog,
)

router = APIRouter(prefix="/exam-sessions", tags=["exam-sessions"])


@router.post("")
def create_exam_session(
    payload: ExamSessionCreate,
    db: Annotated[Session, Depends(get_db)],
    actor: Annotated[User, Depends(require_roles("Super Admin", "Admin"))],
) -> dict:
    row = ExamSession(**payload.model_dump(), created_by=actor.username)
    db.add(row)
    log_event(db, actor_username=actor.username, action="EXAM_SESSION_CREATED", target=payload.course_code)
    db.commit()
    db.refresh(row)
    return {"ok": True, "exam_session": _session_dict(row)}


@router.get("")
def list_exam_sessions(
    db: Annotated[Session, Depends(get_db)],
    _: Annotated[User, Depends(require_roles("Super Admin", "Admin", "Invigilator"))],
) -> dict:
    rows = db.query(ExamSession).order_by(ExamSession.exam_date.desc(), ExamSession.start_time.desc()).all()
    return {"ok": True, "exam_sessions": [_session_dict(row) for row in rows]}


@router.get("/active")
def active_exam_sessions(
    db: Annotated[Session, Depends(get_db)],
    _: Annotated[User, Depends(require_roles("Super Admin", "Admin", "Invigilator"))],
) -> dict:
    rows = db.query(ExamSession).filter(ExamSession.status == "active").order_by(ExamSession.course_code).all()
    return {"ok": True, "exam_sessions": [_session_dict(row) for row in rows]}


@router.get("/assigned-to-me")
def assigned_exam_sessions(
    db: Annotated[Session, Depends(get_db)],
    actor: Annotated[User, Depends(require_roles("Super Admin", "Admin", "Invigilator"))],
) -> dict:
    active = db.query(ExamSession).filter(ExamSession.status == "active")
    assignments_exist = db.query(ExamSessionInvigilator.id).first() is not None
    if actor.role == "Invigilator" and assignments_exist:
        active = active.join(ExamSessionInvigilator).filter(
            ExamSessionInvigilator.invigilator_user_id == actor.id
        )
    rows = active.order_by(ExamSession.course_code).all()
    return {
        "ok": True,
        "assignment_filter_active": actor.role == "Invigilator" and assignments_exist,
        "exam_sessions": [_session_dict(row) for row in rows],
    }


@router.get("/{session_id}")
def get_exam_session(
    session_id: int,
    db: Annotated[Session, Depends(get_db)],
    _: Annotated[User, Depends(require_roles("Super Admin", "Admin", "Invigilator"))],
) -> dict:
    return {"ok": True, "exam_session": _session_dict(_session_or_404(db, session_id))}


@router.put("/{session_id}")
def update_exam_session(
    session_id: int,
    payload: ExamSessionUpdate,
    db: Annotated[Session, Depends(get_db)],
    actor: Annotated[User, Depends(require_roles("Super Admin", "Admin"))],
) -> dict:
    row = _session_or_404(db, session_id)
    for key, value in payload.model_dump(exclude_none=True).items():
        setattr(row, key, value)
    row.updated_at = datetime.utcnow()
    log_event(db, actor_username=actor.username, action="EXAM_SESSION_UPDATED", target=str(session_id))
    db.commit()
    db.refresh(row)
    return {"ok": True, "exam_session": _session_dict(row)}


@router.post("/{session_id}/activate")
def activate_exam_session(
    session_id: int,
    db: Annotated[Session, Depends(get_db)],
    actor: Annotated[User, Depends(require_roles("Super Admin", "Admin"))],
) -> dict:
    row = _session_or_404(db, session_id)
    row.status = "active"
    row.updated_at = datetime.utcnow()
    log_event(db, actor_username=actor.username, action="EXAM_SESSION_ACTIVATED", target=str(session_id))
    db.commit()
    return {"ok": True, "exam_session": _session_dict(row)}


@router.post("/{session_id}/complete")
def complete_exam_session(
    session_id: int,
    db: Annotated[Session, Depends(get_db)],
    actor: Annotated[User, Depends(require_roles("Super Admin", "Admin"))],
) -> dict:
    row = _session_or_404(db, session_id)
    row.status = "completed"
    row.updated_at = datetime.utcnow()
    log_event(db, actor_username=actor.username, action="EXAM_SESSION_COMPLETED", target=str(session_id))
    db.commit()
    return {"ok": True, "exam_session": _session_dict(row)}


@router.post("/{session_id}/assign-invigilator")
def assign_invigilator(
    session_id: int,
    payload: InvigilatorAssignmentIn,
    db: Annotated[Session, Depends(get_db)],
    actor: Annotated[User, Depends(require_roles("Super Admin", "Admin"))],
) -> dict:
    _session_or_404(db, session_id)
    invigilator = db.query(User).filter(
        User.id == payload.invigilator_user_id,
        User.role == "Invigilator",
        User.active.is_(True),
        User.account_status == "approved",
    ).first()
    if invigilator is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Approved invigilator not found")
    row = db.query(ExamSessionInvigilator).filter(
        ExamSessionInvigilator.exam_session_id == session_id,
        ExamSessionInvigilator.invigilator_user_id == invigilator.id,
    ).first()
    if row is None:
        row = ExamSessionInvigilator(
            exam_session_id=session_id,
            invigilator_user_id=invigilator.id,
            assigned_by=actor.username,
        )
        db.add(row)
    row.role_in_session = payload.role_in_session
    log_event(db, actor_username=actor.username, action="INVIGILATOR_ASSIGNED", target=invigilator.username)
    db.commit()
    db.refresh(row)
    return {"ok": True, "assignment": _assignment_dict(row)}


@router.get("/{session_id}/invigilators")
def list_session_invigilators(
    session_id: int,
    db: Annotated[Session, Depends(get_db)],
    _: Annotated[User, Depends(require_roles("Super Admin", "Admin", "Invigilator"))],
) -> dict:
    _session_or_404(db, session_id)
    rows = db.query(ExamSessionInvigilator).filter(
        ExamSessionInvigilator.exam_session_id == session_id
    ).all()
    return {"ok": True, "invigilators": [_assignment_dict(row) for row in rows]}


@router.post("/{session_id}/eligible-students")
def add_eligible_student(
    session_id: int,
    payload: EligibleStudentAdd,
    db: Annotated[Session, Depends(get_db)],
    actor: Annotated[User, Depends(require_roles("Super Admin", "Admin"))],
) -> dict:
    _session_or_404(db, session_id)
    student = db.query(Student).filter(Student.id == payload.student_id).first()
    if student is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Student not found")
    row = db.query(ExamSessionStudent).filter(
        ExamSessionStudent.exam_session_id == session_id,
        ExamSessionStudent.student_id == payload.student_id,
    ).first()
    if row is None:
        row = ExamSessionStudent(exam_session_id=session_id, student_id=payload.student_id)
        db.add(row)
    row.eligibility_type = payload.eligibility_type
    row.eligibility_status = payload.eligibility_status
    row.notes = payload.notes
    row.updated_at = datetime.utcnow()
    log_event(db, actor_username=actor.username, action="EXAM_ELIGIBILITY_UPDATED", target=str(payload.student_id))
    db.commit()
    db.refresh(row)
    return {"ok": True, "eligible_student": _eligibility_dict(row)}


@router.get("/{session_id}/eligible-students")
def list_eligible_students(
    session_id: int,
    db: Annotated[Session, Depends(get_db)],
    _: Annotated[User, Depends(require_roles("Super Admin", "Admin", "Invigilator"))],
) -> dict:
    _session_or_404(db, session_id)
    rows = db.query(ExamSessionStudent).filter(ExamSessionStudent.exam_session_id == session_id).all()
    return {"ok": True, "eligible_students": [_eligibility_dict(row) for row in rows]}


@router.post("/{session_id}/eligible-students/import")
async def import_eligible_students(
    session_id: int,
    file: UploadFile,
    db: Annotated[Session, Depends(get_db)],
    actor: Annotated[User, Depends(require_roles("Super Admin", "Admin"))],
) -> dict:
    session = _session_or_404(db, session_id)
    filename = file.filename or "eligible-students.csv"
    content = await file.read()
    rows = _read_import_rows(filename, content)
    review: list[dict] = []
    counts = {
        "total_rows": len(rows),
        "linked_count": 0,
        "already_added_count": 0,
        "unmatched_count": 0,
        "no_face_count": 0,
        "duplicate_count": 0,
        "invalid_count": 0,
    }
    seen_numbers: set[str] = set()

    for raw in rows:
        student_number = str(raw.get("student_number") or "").strip().upper()
        full_name = str(raw.get("full_name") or "").strip()
        eligibility_type = str(raw.get("eligibility_type") or "regular").strip().lower()
        notes = str(raw.get("notes") or "").strip()
        result = {
            "student_number": student_number,
            "full_name": full_name,
            "issue": "",
            "suggested_action": "",
        }
        if not student_number:
            result.update(issue="Invalid student number", suggested_action="Correct student number")
            counts["invalid_count"] += 1
            review.append(result)
            continue
        if student_number in seen_numbers:
            result.update(issue="Duplicate row", suggested_action="Ignore")
            counts["duplicate_count"] += 1
            review.append(result)
            continue
        seen_numbers.add(student_number)
        if eligibility_type not in {
            "regular", "repeat", "deferred", "supplementary", "manual_override"
        }:
            result.update(issue="Invalid eligibility type", suggested_action="Correct eligibility type")
            counts["invalid_count"] += 1
            review.append(result)
            continue
        student_hash = hash_student_identifier(student_number)
        student = db.query(Student).filter(Student.student_number_hash == student_hash).first()
        if student is None:
            result.update(
                issue="Student not found in biometric database",
                suggested_action="Register face first or correct student number",
            )
            counts["unmatched_count"] += 1
            review.append(result)
            continue
        if not _student_has_face(student):
            result.update(
                full_name=student.full_name,
                issue="Student exists but face not enrolled",
                suggested_action="Register face first",
            )
            counts["no_face_count"] += 1
            review.append(result)
            continue
        existing = db.query(ExamSessionStudent).filter(
            ExamSessionStudent.exam_session_id == session_id,
            ExamSessionStudent.student_id == student.id,
        ).first()
        if existing is not None:
            result.update(
                full_name=student.full_name,
                issue="Already linked to session",
                suggested_action="Ignore",
            )
            counts["already_added_count"] += 1
            review.append(result)
            continue
        db.add(
            ExamSessionStudent(
                exam_session_id=session_id,
                student_id=student.id,
                eligibility_type=eligibility_type,
                eligibility_status="eligible",
                notes=notes or f"Imported from {filename}.",
            )
        )
        counts["linked_count"] += 1

    audit = ExamImportAudit(
        exam_session_id=session_id,
        imported_by=actor.username,
        filename=filename,
        review_json=json.dumps(review),
        **{key: counts[key] for key in (
            "total_rows", "linked_count", "unmatched_count", "no_face_count",
            "duplicate_count", "invalid_count",
        )},
    )
    db.add(audit)
    log_event(
        db,
        actor_username=actor.username,
        action="EXAM_ELIGIBILITY_IMPORTED",
        target=session.course_code,
        metadata={"filename": filename, **counts},
    )
    db.commit()
    return {"ok": True, "filename": filename, **counts, "review": review}


@router.get("/{session_id}/eligible-students/imports")
def list_eligibility_imports(
    session_id: int,
    db: Annotated[Session, Depends(get_db)],
    _: Annotated[User, Depends(require_roles("Super Admin", "Admin", "Invigilator"))],
) -> dict:
    _session_or_404(db, session_id)
    rows = db.query(ExamImportAudit).filter(
        ExamImportAudit.exam_session_id == session_id
    ).order_by(ExamImportAudit.created_at.desc()).all()
    return {"ok": True, "imports": [_import_audit_dict(row) for row in rows]}


@router.post("/{session_id}/eligible-students/from-cohort")
def add_matching_cohort(
    session_id: int,
    db: Annotated[Session, Depends(get_db)],
    actor: Annotated[User, Depends(require_roles("Super Admin", "Admin"))],
) -> dict:
    session = _session_or_404(db, session_id)
    query = db.query(Student).filter(Student.active.is_(True), Student.status == "active")
    if session.program.strip():
        query = query.filter(Student.program.ilike(session.program.strip()))
    if session.level.strip():
        query = query.filter(Student.level.ilike(session.level.strip()))

    added = 0
    for student in query.all():
        if not _student_has_face(student):
            continue
        existing = db.query(ExamSessionStudent).filter(
            ExamSessionStudent.exam_session_id == session_id,
            ExamSessionStudent.student_id == student.id,
        ).first()
        if existing is not None:
            continue
        db.add(
            ExamSessionStudent(
                exam_session_id=session_id,
                student_id=student.id,
                eligibility_type="regular",
                eligibility_status="eligible",
                notes="Added from matching program and level cohort.",
            )
        )
        added += 1
    log_event(
        db,
        actor_username=actor.username,
        action="EXAM_COHORT_ADDED",
        target=f"{session.course_code}:{added}",
    )
    db.commit()
    return {
        "ok": True,
        "added": added,
        "message": (
            f"Added {added} active {session.program} Level {session.level} student(s). "
            "Add repeat, deferred, and supplementary students as exceptions."
        ),
    }


@router.delete("/{session_id}/eligible-students/{student_id}")
def remove_eligible_student(
    session_id: int,
    student_id: int,
    db: Annotated[Session, Depends(get_db)],
    actor: Annotated[User, Depends(require_roles("Super Admin", "Admin"))],
) -> dict:
    deleted = db.query(ExamSessionStudent).filter(
        ExamSessionStudent.exam_session_id == session_id,
        ExamSessionStudent.student_id == student_id,
    ).delete(synchronize_session=False)
    log_event(db, actor_username=actor.username, action="EXAM_ELIGIBILITY_REMOVED", target=str(student_id))
    db.commit()
    if not deleted:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Eligible student link not found")
    return {"ok": True, "message": "Student removed from exam session"}


@router.post("/{session_id}/verify")
def evaluate_exam_entry(
    session_id: int,
    payload: ExamEntryEvaluateIn,
    db: Annotated[Session, Depends(get_db)],
    actor: Annotated[User, Depends(require_roles("Super Admin", "Admin", "Invigilator"))],
) -> dict:
    session = _session_or_404(db, session_id)
    decision = "DENIED"
    reason = ""
    student = db.query(Student).filter(Student.id == payload.detected_student_id).first() if payload.detected_student_id else None
    eligibility = db.query(ExamSessionStudent).filter(
        ExamSessionStudent.exam_session_id == session_id,
        ExamSessionStudent.student_id == payload.detected_student_id,
    ).first() if payload.detected_student_id else None
    session_has_assignments = db.query(ExamSessionInvigilator.id).filter(
        ExamSessionInvigilator.exam_session_id == session_id
    ).first() is not None
    actor_is_assigned = db.query(ExamSessionInvigilator.id).filter(
        ExamSessionInvigilator.exam_session_id == session_id,
        ExamSessionInvigilator.invigilator_user_id == actor.id,
    ).first() is not None

    if actor.role == "Invigilator" and session_has_assignments and not actor_is_assigned:
        reason = "Invigilator is not assigned to the selected exam session."
    elif session.status != "active":
        reason = "No active exam session selected."
    elif not payload.liveness_passed:
        reason = "Liveness failed."
    elif payload.match_score > payload.match_threshold:
        reason = "Match score below threshold."
    elif payload.confidence_gap < payload.minimum_confidence_gap:
        reason = "Similarity gap too small / ambiguous identity."
    elif not payload.identity_matched or student is None:
        reason = "Face not recognized."
    elif not _student_has_face(student):
        reason = "Cannot verify until face enrollment is completed."
    elif student.status != "active" or not student.active:
        reason = "Student inactive or suspended."
    elif eligibility is None:
        reason = "Student is registered in the system but not eligible for this exam session."
    elif eligibility.eligibility_status != "eligible":
        reason = "Student blocked from this exam session."
    elif eligibility.attendance_status == "verified" and not payload.admin_override:
        decision = "ALREADY_VERIFIED"
        reason = _already_verified_reason(eligibility)
    elif payload.admin_override and actor.role not in {"Super Admin", "Admin"}:
        reason = "Only an Admin or Super Admin may override a previous verification."
    elif payload.admin_override and not payload.override_reason.strip():
        reason = "Admin override requires a reason."
    else:
        now = datetime.utcnow()
        updated = db.execute(
            update(ExamSessionStudent)
            .where(
                ExamSessionStudent.id == eligibility.id,
                ExamSessionStudent.attendance_status != "verified",
            )
            .values(
                attendance_status="verified",
                verified_at=now,
                verified_by=actor.username,
                verified_device_id=payload.device_id,
                updated_at=now,
            )
        ).rowcount
        if updated:
            decision = "VERIFIED"
            reason = "Identity, liveness, and exam-session eligibility confirmed."
            eligibility.attendance_status = "verified"
            eligibility.verified_at = now
            eligibility.verified_by = actor.username
            eligibility.verified_device_id = payload.device_id
            if payload.admin_override:
                eligibility.eligibility_type = "manual_override"
                eligibility.notes = payload.override_reason.strip()
        else:
            db.refresh(eligibility)
            decision = "ALREADY_VERIFIED"
            reason = _already_verified_reason(eligibility)

    if eligibility is not None and decision not in {"VERIFIED", "ALREADY_VERIFIED"}:
        eligibility.attendance_status = "denied"
        eligibility.updated_at = datetime.utcnow()

    other_session_activity = False
    if student is not None:
        other_session_activity = db.query(VerificationLog.id).filter(
            VerificationLog.student_id == student.id,
            VerificationLog.exam_session_id != session.id,
            VerificationLog.decision == "VERIFIED",
        ).first() is not None
    device = _touch_device(db, payload, session)

    log = VerificationLog(
        student_id=student.id if student else None,
        exam_session_id=session.id,
        student_number_mask=student.student_number_mask if student else "",
        full_name=student.full_name if student else "Unknown face",
        program=student.program if student else "",
        status=decision,
        decision=decision,
        reason=reason,
        confidence=max(0.0, 1.0 - payload.match_score),
        confidence_gap=payload.confidence_gap,
        liveness_score=1.0 if payload.liveness_passed else 0.0,
        liveness_passed=payload.liveness_passed,
        eligibility_type=eligibility.eligibility_type if eligibility else "",
        verified_by=actor.username,
        device_type=payload.device_type,
        device_id=payload.device_id,
        venue=session.venue,
        metadata_json=json.dumps(
            {
                "admin_override": payload.admin_override,
                "other_session_activity": other_session_activity,
                "device_name": device.device_name if device else payload.device_name,
            },
            sort_keys=True,
        ),
    )
    db.add(log)
    log_event(db, actor_username=actor.username, action="EXAM_ENTRY_DECISION", target=str(session_id), metadata={"decision": decision, "reason": reason})
    db.commit()
    return _decision_dict(
        decision, reason, session, student, eligibility, payload,
        other_session_activity=other_session_activity,
    )


@router.get("/{session_id}/logs")
def exam_session_logs(
    session_id: int,
    db: Annotated[Session, Depends(get_db)],
    _: Annotated[User, Depends(require_roles("Super Admin", "Admin", "Invigilator"))],
) -> dict:
    rows = db.query(VerificationLog).filter(VerificationLog.exam_session_id == session_id).order_by(VerificationLog.created_at.desc()).all()
    return {"ok": True, "logs": [_log_dict(row) for row in rows]}


def _session_or_404(db: Session, session_id: int) -> ExamSession:
    row = db.query(ExamSession).filter(ExamSession.id == session_id).first()
    if row is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Exam session not found")
    return row


def _session_dict(row: ExamSession) -> dict:
    return {
        "id": row.id, "course_code": row.course_code, "course_name": row.course_name,
        "program": row.program, "level": row.level, "exam_date": row.exam_date,
        "start_time": row.start_time, "end_time": row.end_time, "venue": row.venue,
        "status": row.status, "created_by": row.created_by,
        "created_at": row.created_at.isoformat(), "updated_at": row.updated_at.isoformat(),
    }


def _eligibility_dict(row: ExamSessionStudent) -> dict:
    student = row.student
    return {
        "id": row.id, "exam_session_id": row.exam_session_id, "student_id": row.student_id,
        "student_number_mask": student.student_number_mask, "student_name": student.full_name,
        "program": student.program, "level": student.level, "student_status": student.status,
        "biometric_status": "face_enrolled" if _student_has_face(student) else "no_face",
        "eligibility_type": row.eligibility_type, "eligibility_status": row.eligibility_status,
        "attendance_status": row.attendance_status,
        "verified_at": row.verified_at.isoformat() if row.verified_at else None,
        "verified_by": row.verified_by, "verified_device_id": row.verified_device_id,
        "notes": row.notes,
    }


def _student_has_face(student: Student) -> bool:
    from backend.security.data_encryption import decrypt_json

    try:
        profile = decrypt_json(student.biometric_profile_json or "{}")
    except Exception:
        return False
    signature = profile.get("signature")
    return bool(signature and isinstance(signature, list))


def _read_import_rows(filename: str, content: bytes) -> list[dict]:
    suffix = filename.lower().rsplit(".", 1)[-1] if "." in filename else ""
    if suffix == "csv":
        text = content.decode("utf-8-sig")
        return [
            {str(key).strip().lower(): value for key, value in row.items()}
            for row in csv.DictReader(StringIO(text))
        ]
    if suffix == "xlsx":
        workbook = load_workbook(BytesIO(content), read_only=True, data_only=True)
        sheet = workbook.active
        values = sheet.iter_rows(values_only=True)
        headers = [str(value or "").strip().lower() for value in next(values, ())]
        return [
            {headers[index]: value for index, value in enumerate(row) if index < len(headers)}
            for row in values
            if any(value is not None and str(value).strip() for value in row)
        ]
    raise HTTPException(
        status_code=status.HTTP_400_BAD_REQUEST,
        detail="Only CSV and XLSX eligibility files are supported.",
    )


def _import_audit_dict(row: ExamImportAudit) -> dict:
    return {
        "id": row.id,
        "exam_session_id": row.exam_session_id,
        "imported_by": row.imported_by,
        "filename": row.filename,
        "total_rows": row.total_rows,
        "linked_count": row.linked_count,
        "unmatched_count": row.unmatched_count,
        "no_face_count": row.no_face_count,
        "duplicate_count": row.duplicate_count,
        "invalid_count": row.invalid_count,
        "review": json.loads(row.review_json or "[]"),
        "created_at": row.created_at.isoformat(),
    }


def _decision_dict(decision: str, reason: str, session: ExamSession, student: Student | None, eligibility: ExamSessionStudent | None, payload: ExamEntryEvaluateIn, other_session_activity: bool = False) -> dict:
    return {
        "ok": True, "decision": decision, "student_id": student.id if student else None,
        "student_name": student.full_name if student else None,
        "student_number": student.student_number_mask if student else None,
        "match_score": payload.match_score, "confidence_gap": payload.confidence_gap,
        "liveness_passed": payload.liveness_passed, "reason": reason,
        "eligibility_type": eligibility.eligibility_type if eligibility else None,
        "exam_session_id": session.id,
        "verified_at": eligibility.verified_at.isoformat() if eligibility and eligibility.verified_at else None,
        "verified_by": eligibility.verified_by if eligibility else None,
        "verified_device_id": eligibility.verified_device_id if eligibility else None,
        "other_session_activity": other_session_activity,
    }


def _log_dict(row: VerificationLog) -> dict:
    return {
        "id": row.id, "timestamp": row.created_at.isoformat(),
        "exam_session_id": row.exam_session_id, "detected_student_id": row.student_id,
        "detected_student_name": row.full_name, "decision": row.decision or row.status,
        "reason": row.reason, "match_score": 1.0 - row.confidence,
        "confidence_gap": row.confidence_gap, "liveness_passed": row.liveness_passed,
        "eligibility_type": row.eligibility_type, "verified_by": row.verified_by,
        "device_type": row.device_type, "venue": row.venue,
    }


def _assignment_dict(row: ExamSessionInvigilator) -> dict:
    return {
        "id": row.id,
        "exam_session_id": row.exam_session_id,
        "invigilator_user_id": row.invigilator_user_id,
        "invigilator_username": row.invigilator.username,
        "invigilator_name": row.invigilator.full_name,
        "assigned_by": row.assigned_by,
        "assigned_at": row.assigned_at.isoformat(),
        "role_in_session": row.role_in_session,
    }


def _already_verified_reason(row: ExamSessionStudent) -> str:
    return (
        f"Student was already verified at {row.verified_at} by "
        f"{row.verified_by or 'another invigilator'} on "
        f"{row.verified_device_id or 'another device'}."
    )


def _touch_device(db: Session, payload: ExamEntryEvaluateIn, session: ExamSession) -> Device | None:
    if not payload.device_id.strip():
        return None
    row = db.query(Device).filter(Device.device_id == payload.device_id).first()
    if row is None:
        row = Device(device_id=payload.device_id)
        db.add(row)
    row.device_name = payload.device_name.strip() or payload.device_id
    row.device_type = payload.device_type
    row.assigned_room = session.venue
    row.last_seen_at = datetime.utcnow()
    return row
